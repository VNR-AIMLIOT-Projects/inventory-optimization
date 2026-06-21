import io
import pandas as pd
from fpdf import FPDF
from datetime import datetime
from typing import List, Dict, Any

class ReportPDF(FPDF):
    def header(self):
        # Premium dark header bar
        self.set_fill_color(15, 23, 42) # Slate 900
        self.rect(0, 0, 210, 35, 'F')
        
        self.set_y(10)
        self.set_text_color(255, 255, 255)
        self.set_font("helvetica", "B", 20)
        self.cell(0, 8, "REPLENIX INTELLIGENCE", border=False, align="C", new_x="LMARGIN", new_y="NEXT")
        
        self.set_font("helvetica", "", 10)
        self.set_text_color(148, 163, 184) # Slate 400
        self.cell(0, 8, f"Inventory Optimization Report - {datetime.now().strftime('%b %d, %Y %H:%M')}", border=False, align="C", new_x="LMARGIN", new_y="NEXT")
        
        self.ln(15)
        self.set_text_color(15, 23, 42) # Reset to dark text

    def footer(self):
        self.set_y(-15)
        self.set_text_color(148, 163, 184)
        self.set_font("helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

def generate_excel_report(sku: str, metrics: Dict[str, Any], history: List[Dict[str, Any]]) -> io.BytesIO:
    """Generates a beautifully formatted multi-sheet Excel workbook."""
    output = io.BytesIO()
    
    df = pd.DataFrame(history)
    if not df.empty:
        df = df.rename(columns={
            "day": "Day", "date": "Date", "demand": "Demand",
            "inventory": "Inventory", "rl_action": "RL Action",
            "human_action": "Human Action", "final_action": "Final Action",
            "reward": "Reward", "pipeline": "Pipeline"
        })

    summary_data = {
        "Metric": [
            "SKU ID", "Total Simulation Days", "Cumulative RL Reward",
            "Gross Revenue", "Total Operating Cost", "Net Profit",
            "Stockout Events", "Average Inventory Level"
        ],
        "Value": [
            sku,
            metrics.get("total_days", 0),
            f"{metrics.get('cumulative_reward', 0):.2f}",
            f"${metrics.get('total_revenue', 0):.2f}",
            f"${metrics.get('total_cost', 0):.2f}",
            f"${(metrics.get('total_revenue', 0) - metrics.get('total_cost', 0)):.2f}",
            metrics.get("stockout_days", 0),
            f"{metrics.get('avg_inventory', 0):.2f}"
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)
        if not df.empty:
            df.to_excel(writer, sheet_name="Simulation Data", index=False)
            
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            header_font = Font(color="F8FAFC", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            thin_border = Border(left=Side(style='thin', color='E2E8F0'), 
                               right=Side(style='thin', color='E2E8F0'), 
                               top=Side(style='thin', color='E2E8F0'), 
                               bottom=Side(style='thin', color='E2E8F0'))

            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                
                # Auto-adjust column widths & add borders
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        cell.border = thin_border
                        if sheet_name == "Executive Summary" and cell.column == 1:
                            cell.alignment = align_left
                            cell.font = Font(bold=True, color="334155")
                        elif cell.row > 1:
                            cell.alignment = align_center
                            
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 4
        except ImportError:
            pass # Openpyxl not available for styling

    output.seek(0)
    return output


def generate_pdf_report(sku: str, metrics: Dict[str, Any], history: List[Dict[str, Any]]) -> io.BytesIO:
    """Generates a premium formatted PDF document."""
    pdf = ReportPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # --- Executive Summary Section ---
    pdf.set_font("helvetica", "B", 16)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 10, f"Executive Summary: {sku}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_draw_color(226, 232, 240)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)
    
    net_profit = metrics.get('total_revenue', 0) - metrics.get('total_cost', 0)
    
    # 2x2 Grid for key metrics
    pdf.set_fill_color(248, 250, 252) # Slate 50
    pdf.set_draw_color(203, 213, 225) # Slate 300
    
    def metric_box(x, y, title, val, color=(15, 23, 42)):
        pdf.set_xy(x, y)
        pdf.rect(x, y, 90, 20, 'DF')
        pdf.set_xy(x+5, y+4)
        pdf.set_font("helvetica", "B", 8)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(80, 5, title.upper(), align="L")
        pdf.set_xy(x+5, y+9)
        pdf.set_font("helvetica", "B", 14)
        pdf.set_text_color(*color)
        pdf.cell(80, 8, val, align="L")

    y_start = pdf.get_y()
    metric_box(10, y_start, "Net Profit", f"${net_profit:,.2f}", color=(16, 185, 129)) # Emerald 500
    metric_box(110, y_start, "Gross Revenue", f"${metrics.get('total_revenue', 0):,.2f}")
    
    y_start += 25
    metric_box(10, y_start, "Total Operating Cost", f"${metrics.get('total_cost', 0):,.2f}", color=(239, 68, 68)) # Red 500
    metric_box(110, y_start, "Stockout Days", str(metrics.get('stockout_days', 0)))
    
    y_start += 25
    metric_box(10, y_start, "Cumulative RL Reward", f"{metrics.get('cumulative_reward', 0):,.2f}")
    metric_box(110, y_start, "Avg Inventory Level", f"{metrics.get('avg_inventory', 0):,.2f} units")
    
    pdf.set_y(y_start + 30)
    
    # --- Data Table Section ---
    pdf.set_font("helvetica", "B", 16)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 10, "Day-by-Day Simulation Log", new_x="LMARGIN", new_y="NEXT")
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)
    
    # Table Header
    pdf.set_fill_color(15, 23, 42)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 9)
    col_widths = [15, 25, 20, 20, 25, 25, 20]
    headers = ["Day", "Date", "Demand", "Inv", "RL Action", "Final Act", "Reward"]
    
    for width, header in zip(col_widths, headers):
        pdf.cell(width, 10, header, border=0, align="C", fill=True)
    pdf.ln()
    
    # Table Rows
    pdf.set_font("helvetica", "", 9)
    pdf.set_text_color(51, 65, 85)
    fill = False
    
    for row in history:
        if fill:
            pdf.set_fill_color(248, 250, 252)
        else:
            pdf.set_fill_color(255, 255, 255)
            
        pdf.cell(col_widths[0], 8, str(row.get("day", "")), border="B", align="C", fill=True)
        pdf.cell(col_widths[1], 8, str(row.get("date", "")), border="B", align="C", fill=True)
        pdf.cell(col_widths[2], 8, str(row.get("demand", "")), border="B", align="C", fill=True)
        pdf.cell(col_widths[3], 8, f"{row.get('inventory', 0):.1f}", border="B", align="C", fill=True)
        pdf.cell(col_widths[4], 8, str(row.get("rl_action", "")), border="B", align="C", fill=True)
        
        # Color final action
        final_act = str(row.get("final_action", ""))
        if final_act != str(row.get("rl_action", "")):
            pdf.set_text_color(217, 119, 6) # Amber
        pdf.cell(col_widths[5], 8, final_act, border="B", align="C", fill=True)
        pdf.set_text_color(51, 65, 85) # Reset
        
        reward = row.get('reward', 0)
        if reward < 0:
            pdf.set_text_color(239, 68, 68)
        else:
            pdf.set_text_color(16, 185, 129)
        pdf.cell(col_widths[6], 8, f"{reward:.1f}", border="B", align="C", fill=True)
        pdf.set_text_color(51, 65, 85) # Reset
        
        pdf.ln()
        fill = not fill
        
    output = io.BytesIO(pdf.output())
    output.seek(0)
    return output
